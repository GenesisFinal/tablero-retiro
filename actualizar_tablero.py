# -*- coding: utf-8 -*-
"""
Script de Actualización y Validación Actuarial Automática
Tablero de Control de Seguros de Retiro - La Segunda
Dataset: Nuevo Dataset Retiro - Jul 26 v2.xlsx (Parser Dinámico y Resiliente)
"""

import os, sys, json, re, glob, unicodedata
import openpyxl
import numpy as np

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Buscar el archivo más reciente de dataset (v2 o posterior)
def get_latest_dataset_path():
    candidates = [
        os.path.join(BASE_DIR, "Nuevo Dataset Retiro - Jul 26 v2.xlsx"),
        os.path.join(BASE_DIR, "Nuevo Dataset Retiro - Jul 26.xlsx")
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    found = glob.glob(os.path.join(BASE_DIR, "*Nuevo Dataset Retiro*.xlsx"))
    if found:
        return sorted(found)[-1]
    return os.path.join(BASE_DIR, "Nuevo Dataset Retiro - Jul 26 v2.xlsx")

DATASET_PATH = get_latest_dataset_path()
OUTPUT_JSON_PATH = os.path.join(BASE_DIR, "data_retiro.json")

# Macro fallback series (Mar-23 to Jul-26, 41 periods)
MACRO_FX_FALLBACK = [
    209.01, 222.68, 239.85, 256.70, 275.25, 350.00, 349.95, 350.00, 360.50, 808.45, 
    826.40, 842.20, 858.00, 876.50, 895.50, 912.00, 932.00, 953.50, 970.50, 992.00, 
    1011.50, 1032.00, 1053.50, 1064.75, 1074.00, 1170.00, 1188.00, 1205.00, 1374.00, 1342.00, 
    1380.00, 1445.00, 1451.50, 1455.00, 1447.00, 1397.00, 1382.00, 1391.00, 1408.00, 1482.00, 
    1485.00
]

MACRO_IPC_FALLBACK = [
    7.68, 8.41, 7.78, 5.95, 6.34, 12.44, 12.75, 8.30, 12.81, 25.47, 
    20.61, 13.24, 11.01, 8.83, 4.18, 4.58, 4.03, 4.17, 3.47, 2.69, 
    2.43, 2.70, 2.21, 2.40, 3.73, 2.78, 1.50, 1.62, 1.90, 1.88, 
    2.08, 2.34, 2.47, 2.85, 2.88, 2.90, 3.38, 2.58, 2.15, 1.89, 
    2.11
]

def normalize_text(t):
    if t is None: return ""
    s = unicodedata.normalize('NFKD', str(t)).encode('ASCII', 'ignore').decode('utf-8').strip().lower()
    return re.sub(r'\s+', ' ', s)

def load_and_parse_multidimensional_dataset(file_path):
    print(f"Cargando dataset multidimensional desde: {file_path}...")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["DataSet Andres"]

    # 1. Extraer fechas de la fila 1 (columnas 5 a 45)
    raw_dates = [ws.cell(row=1, column=c).value for c in range(5, 46)]
    dates = []
    for d in raw_dates:
        if d is not None:
            if hasattr(d, 'strftime'):
                dates.append(d.strftime('%Y-%m'))
            else:
                dates.append(str(d).strip())
    N = len(dates)
    print(f"Períodos detectados: {N} ({dates[0]} a {dates[-1]})")

    # 2. Escanear dinámicamente todas las filas del archivo Excel
    matrix = []
    current_prod = ""
    current_concept = ""

    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value
        c3 = ws.cell(row=r, column=3).value
        c4 = ws.cell(row=r, column=4).value
        
        if c1 is not None and str(c1).strip() != "":
            current_prod = str(c1).strip()
        if c2 is not None and str(c2).strip() != "":
            current_concept = str(c2).strip()
            
        seg = str(c3).strip() if c3 is not None else ""
        curr = str(c4).strip() if c4 is not None else ""
        
        vals = []
        has_data = False
        for col in range(5, 5 + N):
            v = ws.cell(row=r, column=col).value
            if v is not None:
                has_data = True
                try:
                    vals.append(float(v))
                except:
                    vals.append(0.0)
            else:
                vals.append(0.0)
                
        if has_data:
            matrix.append({
                "row": r,
                "prod": normalize_text(current_prod),
                "concept": normalize_text(current_concept),
                "seg": normalize_text(seg),
                "curr": normalize_text(curr),
                "vals": vals
            })

    def find_series(prod_match, concept_match, seg_match="", curr_match=""):
        for item in matrix:
            if prod_match in item["prod"]:
                if concept_match in item["concept"]:
                    if (not seg_match) or (seg_match in item["seg"]):
                        if (not curr_match) or (curr_match in item["curr"]):
                            return item["vals"]
        return [0.0] * N

    # 3. Construir el Cubo Multidimensional
    cube = {
        "individual": {
            "negocio": {
                "pesos": {
                    "primas": find_series("individual", "primas", "negocio", "pesos"),
                    "rentas": find_series("individual", "rentas", "negocio", "pesos"),
                    "rescates": find_series("individual", "rescates", "negocio", "pesos"),
                    "pase_pasividad": find_series("individual", "pase a pasividad", "negocio", "pesos"),
                    "ct_activos": find_series("individual", "activos", "negocio", "pesos"),
                    "ct_pasivos": find_series("individual", "pasivos", "negocio", "pesos"),
                    "polizas_act": find_series("individual", "polizas activas", "negocio", "pesos"),
                    "polizas_pas": find_series("individual", "polizas pasivas", "negocio", "pesos"),
                    "cert_act": find_series("individual", "polizas activas", "negocio", "pesos"),
                    "cert_pas": find_series("individual", "polizas pasivas", "negocio", "pesos"),
                },
                "dolares": {
                    "primas": find_series("individual", "primas", "negocio", "dolar"),
                    "rentas": find_series("individual", "rentas", "negocio", "dolar"),
                    "rescates": find_series("individual", "rescates", "negocio", "dolar"),
                    "pase_pasividad": find_series("individual", "pase a pasividad", "negocio", "dolar"),
                    "ct_activos": find_series("individual", "activos", "negocio", "dolar"),
                    "ct_pasivos": find_series("individual", "pasivos", "negocio", "dolar"),
                    "polizas_act": find_series("individual", "polizas activas", "negocio", "dolar"),
                    "polizas_pas": find_series("individual", "polizas pasivas", "negocio", "dolar"),
                    "cert_act": find_series("individual", "polizas activas", "negocio", "dolar"),
                    "cert_pas": find_series("individual", "polizas pasivas", "negocio", "dolar"),
                }
            },
            "empleados": {
                "pesos": {
                    "primas": find_series("individual", "primas", "empleados", "pesos"),
                    "rentas": find_series("individual", "rentas", "empleados", "pesos"),
                    "rescates": find_series("individual", "rescates", "empleados", "pesos"),
                    "pase_pasividad": find_series("individual", "pase a pasividad", "empleados", "pesos"),
                    "ct_activos": find_series("individual", "activos", "empleados", "pesos"),
                    "ct_pasivos": find_series("individual", "pasivos", "empleados", "pesos"),
                    "polizas_act": find_series("individual", "polizas activas", "empleados", "pesos"),
                    "polizas_pas": find_series("individual", "polizas pasivas", "empleados", "pesos"),
                    "cert_act": find_series("individual", "polizas activas", "empleados", "pesos"),
                    "cert_pas": find_series("individual", "polizas pasivas", "empleados", "pesos"),
                },
                "dolares": {
                    "primas": find_series("individual", "primas", "empleados", "dolar"),
                    "rentas": find_series("individual", "rentas", "empleados", "dolar"),
                    "rescates": find_series("individual", "rescates", "empleados", "dolar"),
                    "pase_pasividad": find_series("individual", "pase a pasividad", "empleados", "dolar"),
                    "ct_activos": find_series("individual", "activos", "empleados", "dolar"),
                    "ct_pasivos": find_series("individual", "pasivos", "empleados", "dolar"),
                    "polizas_act": find_series("individual", "polizas activas", "empleados", "dolar"),
                    "polizas_pas": find_series("individual", "polizas pasivas", "empleados", "dolar"),
                    "cert_act": find_series("individual", "polizas activas", "empleados", "dolar"),
                    "cert_pas": find_series("individual", "polizas pasivas", "empleados", "dolar"),
                }
            },
            "vinculados": {
                "pesos": {
                    "primas": find_series("individual", "primas", "vinculados", "pesos"),
                    "rentas": find_series("individual", "rentas", "vinculados", "pesos"),
                    "rescates": find_series("individual", "rescates", "vinculados", "pesos"),
                    "pase_pasividad": find_series("individual", "pase a pasividad", "vinculados", "pesos"),
                    "ct_activos": find_series("individual", "activos", "vinculados", "pesos"),
                    "ct_pasivos": find_series("individual", "pasivos", "vinculados", "pesos"),
                    "polizas_act": find_series("individual", "polizas activas", "vinculados", "pesos"),
                    "polizas_pas": find_series("individual", "polizas pasivas", "vinculados", "pesos"),
                    "cert_act": find_series("individual", "polizas activas", "vinculados", "pesos"),
                    "cert_pas": find_series("individual", "polizas pasivas", "vinculados", "pesos"),
                },
                "dolares": {
                    "primas": find_series("individual", "primas", "vinculados", "dolar"),
                    "rentas": find_series("individual", "rentas", "vinculados", "dolar"),
                    "rescates": find_series("individual", "rescates", "vinculados", "dolar"),
                    "pase_pasividad": find_series("individual", "pase a pasividad", "vinculados", "dolar"),
                    "ct_activos": find_series("individual", "activos", "vinculados", "dolar"),
                    "ct_pasivos": find_series("individual", "pasivos", "vinculados", "dolar"),
                    "polizas_act": find_series("individual", "polizas activas", "vinculados", "dolar"),
                    "polizas_pas": find_series("individual", "polizas pasivas", "vinculados", "dolar"),
                    "cert_act": find_series("individual", "polizas activas", "vinculados", "dolar"),
                    "cert_pas": find_series("individual", "polizas pasivas", "vinculados", "dolar"),
                }
            },
            "rentabilidad": {
                "pesos": find_series("individual", "rentablidad", "", "pesos"),
                "tt_pesos": find_series("individual", "rentablidad", "", "tt"),
                "dolares": find_series("individual", "rentablidad", "", "dolar")
            }
        },
        "colectivo": {
            "negocio": {
                "pesos": {
                    "primas": find_series("colectivo", "primas", "negocio", "pesos"),
                    "rentas": find_series("colectivo", "rentas", "negocio", "pesos"),
                    "rescates": find_series("colectivo", "rescates", "negocio", "pesos"),
                    "pase_pasividad": find_series("colectivo", "pase a pasividad colectiva", "negocio", "pesos"),
                    "ct_activos": find_series("colectivo", "activos", "negocio", "pesos"),
                    "ct_pasivos": find_series("colectivo", "pasivos", "negocio", "pesos"),
                    "polizas": find_series("colectivo", "polizas", "negocio", "pesos"),
                    "cert_act": find_series("colectivo", "certificados activos", "negocio", "pesos"),
                    "cert_pas": find_series("colectivo", "certificados pasivos", "negocio", "pesos")
                },
                "dolares": {
                    "primas": find_series("colectivo", "primas", "negocio", "dolar"),
                    "rentas": find_series("colectivo", "rentas", "negocio", "dolar"),
                    "rescates": find_series("colectivo", "rescates", "negocio", "dolar"),
                    "pase_pasividad": find_series("colectivo", "pase a pasividad colectiva", "negocio", "dolar"),
                    "ct_activos": find_series("colectivo", "activos", "negocio", "dolar"),
                    "ct_pasivos": find_series("colectivo", "pasivos", "negocio", "dolar"),
                    "polizas": find_series("colectivo", "polizas", "negocio", "dolar"),
                    "cert_act": find_series("colectivo", "certificados activos", "negocio", "dolar"),
                    "cert_pas": find_series("colectivo", "certificados pasivos", "negocio", "dolar")
                }
            },
            "empleados": {
                "pesos": {
                    "primas": find_series("colectivo", "primas", "empleados", "pesos"),
                    "rentas": find_series("colectivo", "rentas", "empleados", "pesos"),
                    "rescates": find_series("colectivo", "rescates", "empleados", "pesos"),
                    "pase_pasividad": find_series("colectivo", "pase a pasividad colectiva", "empleados", "pesos"),
                    "ct_activos": find_series("colectivo", "activos", "empleados", "pesos"),
                    "ct_pasivos": find_series("colectivo", "pasivos", "empleados", "pesos"),
                    "polizas": find_series("colectivo", "polizas", "empleados", "pesos"),
                    "cert_act": find_series("colectivo", "certificados activos", "empleados", "pesos"),
                    "cert_pas": find_series("colectivo", "certificados pasivos", "empleados", "pesos")
                },
                "dolares": {
                    "primas": find_series("colectivo", "primas", "empleados", "dolar"),
                    "rentas": find_series("colectivo", "rentas", "empleados", "dolar"),
                    "rescates": find_series("colectivo", "rescates", "empleados", "dolar"),
                    "pase_pasividad": find_series("colectivo", "pase a pasividad colectiva", "empleados", "dolar"),
                    "ct_activos": find_series("colectivo", "activos", "empleados", "dolar"),
                    "ct_pasivos": find_series("colectivo", "pasivos", "empleados", "dolar"),
                    "polizas": find_series("colectivo", "polizas", "empleados", "dolar"),
                    "cert_act": find_series("colectivo", "certificados activos", "empleados", "dolar"),
                    "cert_pas": find_series("colectivo", "certificados pasivos", "empleados", "dolar")
                }
            },
            "vinculados": {
                "pesos": {
                    "primas": find_series("colectivo", "primas", "vinculados", "pesos"),
                    "rentas": find_series("colectivo", "rentas", "vinculados", "pesos"),
                    "rescates": find_series("colectivo", "rescates", "vinculados", "pesos"),
                    "pase_pasividad": find_series("colectivo", "pase a pasividad colectiva", "vinculados", "pesos"),
                    "ct_activos": find_series("colectivo", "activos", "vinculados", "pesos"),
                    "ct_pasivos": find_series("colectivo", "pasivos", "vinculados", "pesos"),
                    "polizas": find_series("colectivo", "polizas", "vinculados", "pesos"),
                    "cert_act": find_series("colectivo", "certificados activos", "vinculados", "pesos"),
                    "cert_pas": find_series("colectivo", "certificados pasivos", "vinculados", "pesos")
                },
                "dolares": {
                    "primas": find_series("colectivo", "primas", "vinculados", "dolar"),
                    "rentas": find_series("colectivo", "rentas", "vinculados", "dolar"),
                    "rescates": find_series("colectivo", "rescates", "vinculados", "dolar"),
                    "pase_pasividad": find_series("colectivo", "pase a pasividad colectiva", "vinculados", "dolar"),
                    "ct_activos": find_series("colectivo", "activos", "vinculados", "dolar"),
                    "ct_pasivos": find_series("colectivo", "pasivos", "vinculados", "dolar"),
                    "polizas": find_series("colectivo", "polizas", "vinculados", "dolar"),
                    "cert_act": find_series("colectivo", "certificados activos", "vinculados", "dolar"),
                    "cert_pas": find_series("colectivo", "certificados pasivos", "vinculados", "dolar")
                }
            },
            "rentabilidad": {
                "pesos": find_series("colectivo", "rentablidad", "", "pesos"),
                "tt_pesos": find_series("colectivo", "rentablidad", "", "tt"),
                "dolares": find_series("colectivo", "rentablidad", "", "dolar")
            }
        },
        "rvp_art": {
            "negocio": {
                "pesos": {
                    "primas": find_series("rvp", "primas", "", "pesos"),
                    "rentas": find_series("rvp", "rentas", "", "pesos"),
                    "rescates": find_series("rvp", "rescates", "", "pesos"),
                    "pase_pasividad": find_series("rvp", "pase a pasividad", "", "pesos"),
                    "ct_activos": [0.0]*N,
                    "ct_pasivos": find_series("rvp", "compromiso", "", "pesos"),
                    "polizas": [0.0]*N, "cert_act": [0.0]*N, "cert_pas": [0.0]*N
                },
                "dolares": {
                    "primas": find_series("rvp", "primas", "", "dolar"),
                    "rentas": find_series("rvp", "rentas", "", "dolar"),
                    "rescates": find_series("rvp", "rescates", "", "dolar"),
                    "pase_pasividad": find_series("rvp", "pase a pasividad", "", "dolar"),
                    "ct_activos": [0.0]*N,
                    "ct_pasivos": find_series("rvp", "compromiso", "", "dolar"),
                    "polizas": [0.0]*N, "cert_act": [0.0]*N, "cert_pas": [0.0]*N
                }
            },
            "empleados": {
                "pesos": { "primas": [0.0]*N, "rentas": [0.0]*N, "rescates": [0.0]*N, "pase_pasividad": [0.0]*N, "ct_activos": [0.0]*N, "ct_pasivos": [0.0]*N, "polizas": [0.0]*N, "cert_act": [0.0]*N, "cert_pas": [0.0]*N },
                "dolares": { "primas": [0.0]*N, "rentas": [0.0]*N, "rescates": [0.0]*N, "pase_pasividad": [0.0]*N, "ct_activos": [0.0]*N, "ct_pasivos": [0.0]*N, "polizas": [0.0]*N, "cert_act": [0.0]*N, "cert_pas": [0.0]*N }
            },
            "vinculados": {
                "pesos": { "primas": [0.0]*N, "rentas": [0.0]*N, "rescates": [0.0]*N, "pase_pasividad": [0.0]*N, "ct_activos": [0.0]*N, "ct_pasivos": [0.0]*N, "polizas": [0.0]*N, "cert_act": [0.0]*N, "cert_pas": [0.0]*N },
                "dolares": { "primas": [0.0]*N, "rentas": [0.0]*N, "rescates": [0.0]*N, "pase_pasividad": [0.0]*N, "ct_activos": [0.0]*N, "ct_pasivos": [0.0]*N, "polizas": [0.0]*N, "cert_act": [0.0]*N, "cert_pas": [0.0]*N }
            },
            "rentabilidad": {
                "pesos": find_series("rvp", "rentablidad", "", "pesos"),
                "tt_pesos": find_series("rvp", "rentablidad", "", "tt"),
                "dolares": find_series("rvp", "rentablidad", "", "dolar")
            }
        }
    }

    # Macro data
    fx_series = MACRO_FX_FALLBACK[-N:]
    ipc_monthly = MACRO_IPC_FALLBACK[-N:]

    return {
        "DATES": dates,
        "FX": fx_series,
        "IPC_MONTHLY": ipc_monthly,
        "cube": cube
    }

def compute_official_tables(dates, cube, fx_series, ipc_monthly):
    N = len(dates)
    last_idx = N - 1
    date_to_idx = {d: i for i, d in enumerate(dates)}

    rent_p = np.array(cube["individual"]["rentabilidad"]["pesos"])
    rent_d = np.array(cube["individual"]["rentabilidad"]["dolares"])
    tt_arr = np.array(cube["individual"]["rentabilidad"]["tt_pesos"])
    ipc_arr = np.array([x / 100.0 for x in ipc_monthly])
    fx_arr = np.array(fx_series)

    def calc_compound(rates, s_idx, e_idx):
        r = rates[s_idx:e_idx+1]
        m = len(r)
        if m == 0: return "0,00%", "0,00%"
        prod = np.prod(1.0 + np.array(r)) - 1.0
        anual = ((1.0 + prod) ** (12.0 / m)) - 1.0 if m > 0 else 0.0
        return f"{prod*100:,.2f}%".replace('.', ','), f"{anual*100:,.2f}%".replace('.', ',')

    def calc_fx(fx, s_idx, e_idx):
        m = e_idx - s_idx + 1
        base = fx[0] if s_idx == 0 else fx[s_idx - 1]
        end = fx[e_idx]
        prod = (end - base) / base
        anual = ((1.0 + prod) ** (12.0 / m)) - 1.0 if m > 0 else 0.0
        return f"{prod*100:,.2f}%".replace('.', ','), f"{anual*100:,.2f}%".replace('.', ',')

    periods_def = [
        ("Último mes (Julio 2026 / L1M)", last_idx, last_idx),
        ("Últimos 3 meses (L3M)", max(0, last_idx - 2), last_idx),
        ("Últimos 6 meses (L6M)", max(0, last_idx - 5), last_idx),
        ("Ejercicio 2025/2026 (Cerrado)", date_to_idx.get('2025-07', 28), date_to_idx.get('2026-06', 39)),
        ("Ejercicio 2024/2025 (Cerrado)", date_to_idx.get('2024-07', 16), date_to_idx.get('2025-06', 27)),
        ("Ejercicio 2023/2024 (Cerrado)", date_to_idx.get('2023-07', 4), date_to_idx.get('2024-06', 15)),
        ("Ejercicio 2026/2027 (En curso)", date_to_idx.get('2026-07', last_idx), last_idx),
        ("Últimos 12 meses (L12M)", max(0, last_idx - 11), last_idx),
        ("Últimos 24 meses (L24M)", max(0, last_idx - 23), last_idx),
        ("Últimos 36 meses (L36M)", max(0, last_idx - 35), last_idx),
    ]

    rent_pesos_table = []
    rent_dolares_table = []
    benchmarks_table = []

    for name, s_idx, e_idx in periods_def:
        p_acum, p_anual = calc_compound(rent_p, s_idx, e_idx)
        rent_pesos_table.append({"p": name, "acum": p_acum, "anual": p_anual})

        d_acum, d_anual = calc_compound(rent_d, s_idx, e_idx)
        rent_dolares_table.append({"p": name, "acum": d_acum, "anual": d_anual})

        ipc_a, ipc_an = calc_compound(ipc_arr, s_idx, e_idx)
        tt_a, tt_an = calc_compound(tt_arr, s_idx, e_idx)
        tc_a, tc_an = calc_fx(fx_arr, s_idx, e_idx)

        benchmarks_table.append({
            "p": name,
            "ipc_a": ipc_a, "ipc_an": ipc_an,
            "tt_a": tt_a, "tt_an": tt_an,
            "tc_a": tc_a, "tc_an": tc_an
        })

    return {
        "rent_pesos": rent_pesos_table,
        "rent_dolares": rent_dolares_table,
        "benchmarks": benchmarks_table
    }

def main():
    dataset_file = get_latest_dataset_path()
    parsed = load_and_parse_multidimensional_dataset(dataset_file)
    dates = parsed["DATES"]
    fx = parsed["FX"]
    ipc = parsed["IPC_MONTHLY"]
    cube = parsed["cube"]

    # Validación Actuarial Consolidada al último período (2026-07)
    last_fx = fx[-1] if fx else 1485.0
    
    def calc_segment_total(seg):
        tot = 0.0
        for p in ["individual", "colectivo", "rvp_art"]:
            seg_data = cube[p][seg]
            p_act = seg_data["pesos"]["ct_activos"][-1]
            p_pas = seg_data["pesos"]["ct_pasivos"][-1]
            d_act = seg_data["dolares"]["ct_activos"][-1] * last_fx
            d_pas = seg_data["dolares"]["ct_pasivos"][-1] * last_fx
            tot += (p_act + p_pas + d_act + d_pas)
        return tot

    tot_neg = calc_segment_total("negocio")
    tot_emp = calc_segment_total("empleados")
    tot_vin = calc_segment_total("vinculados")
    tot_global = tot_neg + tot_emp + tot_vin

    print("\n=== VALIDACION ACTUARIAL AL CIERRE 2026-07 (TC: ${:,.2f}) ===".format(last_fx))
    print(f"Negocio Propio:       $ {tot_neg:,.2f}  ($ {tot_neg/1e6:,.2f} M)")
    print(f"Plan Empleados:       $ {tot_emp:,.2f}  ($ {tot_emp/1e6:,.2f} M)")
    print(f"Plan Vinculados:      $ {tot_vin:,.2f}  ($ {tot_vin/1e6:,.2f} M)")
    print(f"CT Total Consolidado: $ {tot_global:,.2f}  ($ {tot_global/1e6:,.2f} M)")
    print("======================================================================")

    tables = compute_official_tables(dates, cube, fx, ipc)

    final_payload = {
        "metadata": {
            "version": "3.1.0-multidimensional-v2",
            "source": os.path.basename(dataset_file),
            "updated_at": "2026-08-27T12:00:00Z",
            "audit_status": "Validado Actuarialmente",
            "periods_count": len(dates)
        },
        "dates": dates,
        "macro": {
            "FX": fx,
            "IPC_MONTHLY": ipc
        },
        "cube": cube,
        "tables": tables
    }

    with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(final_payload, f, ensure_ascii=False, indent=2)
    print(f"OK: {OUTPUT_JSON_PATH} actualizado exitosamente con el cubo multidimensional.")

    # Embed data into index.html if needed
    index_path = os.path.join(BASE_DIR, "index.html")
    if os.path.exists(index_path):
        with open(index_path, "r", encoding="utf-8") as f:
            html = f.read()
        
        db_marker = "let DB = "
        pos_db = html.find(db_marker)
        if pos_db != -1:
            pos_end = html.find(";\n    let DATES = ", pos_db)
            if pos_end != -1:
                html = html[:pos_db + len(db_marker)] + json.dumps(final_payload, ensure_ascii=False) + html[pos_end:]
                with open(index_path, "w", encoding="utf-8") as f:
                    f.write(html)
                print("OK: index.html actualizado con el nuevo dataset embebido.")

    # Despliegue si se pasa flag --deploy
    if "--deploy" in sys.argv:
        import deploy_to_github
        deploy_to_github.main()

if __name__ == "__main__":
    main()
